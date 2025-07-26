from fastapi import FastAPI, UploadFile, File
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import shutil

app = FastAPI()

@app.post("/compare")
async def compare_excel(file1: UploadFile = File(...), file2: UploadFile = File(...)):
    path1 = f"/tmp/{file1.filename}"
    path2 = f"/tmp/{file2.filename}"
    output_path = "/tmp/result.xlsx"

    with open(path1, "wb") as buffer:
        shutil.copyfileobj(file1.file, buffer)

    with open(path2, "wb") as buffer:
        shutil.copyfileobj(file2.file, buffer)

    wb1 = load_workbook(path1)
    wb2 = load_workbook(path2)
    ws1 = wb1.active
    ws2 = wb2.active

    fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    for row in range(1, ws1.max_row + 1):
        for col in range(1, ws1.max_column + 1):
            val1 = ws1.cell(row, col).value
            val2 = ws2.cell(row, col).value
            if val1 != val2:
                ws1.cell(row, col).fill = fill

    wb1.save(output_path)
    return FileResponse(output_path, filename="result.xlsx")
